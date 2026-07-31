#!/usr/bin/env python3
"""
Regenerate data/products.txt and data/groups.txt from the source spreadsheets.

    pip install openpyxl
    python3 src/import_sheets.py data/Vendor_Group.xlsx

Vendor_Group.xlsx layout:
  * a "Products" tab  -> Code | Product | Tamil name | Type | Selling margin %
  * one tab per group -> Code | Product | Tamil name
    The tab NAME is the group name. Any product not listed in a group tab
    is treated as "Manual order" by the app itself, so nothing is lost.

products.txt line:  code|English / Tamil|unit|unit weight kg|selling margin %|alias
groups.txt line:    group name|code,code,code
"""
import pathlib
import sys

from openpyxl import load_workbook

ROOT = pathlib.Path(__file__).resolve().parent.parent


def norm(v):
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    s = str(v).strip()
    return s or None


def main(path: str) -> int:
    wb = load_workbook(path, data_only=True)
    if "Products" not in wb.sheetnames:
        print("ERROR: no 'Products' tab found", file=sys.stderr)
        return 1

    # --- products ---
    ws = wb["Products"]
    existing = {}
    old = ROOT / "data" / "products.txt"
    if old.exists():
        for line in old.read_text(encoding="utf-8").splitlines():
            f = line.split("|")
            if f and f[0].strip():
                existing[f[0].strip()] = f

    rows = []
    for r in range(2, ws.max_row + 1):
        code = norm(ws.cell(row=r, column=1).value)
        if not code:
            continue
        eng = norm(ws.cell(row=r, column=2).value) or ""
        tam = norm(ws.cell(row=r, column=3).value) or ""
        unit = norm(ws.cell(row=r, column=4).value) or "kg"
        margin = norm(ws.cell(row=r, column=5).value) or ""
        # unit weight and alias are not in this sheet — keep whatever we already had
        prev = existing.get(code, ["", "", "", "", "", ""])
        weight = prev[3] if len(prev) > 3 else ""
        alias = prev[5] if len(prev) > 5 else ""
        rows.append(f"{code}|{eng} / {tam}|{unit}|{weight}|{margin}|{alias}")
    (ROOT / "data" / "products.txt").write_text("\n".join(rows) + "\n", encoding="utf-8")

    # --- groups ---
    glines = []
    for name in wb.sheetnames:
        if name == "Products":
            continue
        ws = wb[name]
        codes = []
        for r in range(2, ws.max_row + 1):
            c = norm(ws.cell(row=r, column=1).value)
            if c:
                codes.append(c)
        glines.append(f"{name}|{','.join(codes)}")
    (ROOT / "data" / "groups.txt").write_text("\n".join(glines) + "\n", encoding="utf-8")

    grouped = sum(len(l.split("|")[1].split(",")) for l in glines if l.split("|")[1])
    print(f"products: {len(rows)}   groups: {len(glines)}   grouped codes: {grouped}")
    print(f"ungrouped (fall into 'Manual order'): {len(rows) - grouped}")
    print("now run:  python3 src/build.py")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1] if len(sys.argv) > 1 else str(ROOT / "data" / "Vendor_Group.xlsx")))
